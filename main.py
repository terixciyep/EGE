import asyncio
import random
import copy
import os
from typing import Callable, Dict, Any, Awaitable
from aiogram import Bot, Dispatcher, F, BaseMiddleware
from aiogram.types import (
    Message,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    CallbackQuery,
    Update
)
from aiogram.filters import CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN")
XLSX_PATH = os.getenv("XLSX_PATH", "fic.xlsx")
REQUIRED_CHANNEL_ID = os.getenv("REQUIRED_CHANNEL_ID")
CHANNEL_LINK = os.getenv("CHANNEL_LINK")

if REQUIRED_CHANNEL_ID:
    try:
        REQUIRED_CHANNEL_ID = int(REQUIRED_CHANNEL_ID)
    except ValueError:
        raise ValueError(
            f"REQUIRED_CHANNEL_ID должен быть числовым ID канала (например: -1001234567890), "
            f"получено: {REQUIRED_CHANNEL_ID}"
        )

# Валидация обязательных переменных
if not BOT_TOKEN:
    raise ValueError("BOT_TOKEN не найден в .env файле")


# ---------- FSM ----------

class QuizState(StatesGroup):
    choosing_topic = State()
    answering = State()


# ---------- Загрузка Excel ----------

def load_quiz_from_xlsx(path: str) -> dict:
    wb = load_workbook(path)
    data = {}

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        questions = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            question = str(row[0]).strip()

            options = [
                str(row[1]).strip(),
                str(row[2]).strip(),
                str(row[3]).strip(),
                str(row[4]).strip()
            ]

            correct = str(row[5]).strip()

            questions.append({
                "question": question,
                "options": options,
                "correct": correct
            })

        data[sheet] = questions

    return data



QUIZ_DATA = load_quiz_from_xlsx(XLSX_PATH)

# Множество пользователей, которые ранее работали с ботом
active_users = set()


# ---------- Проверка подписки ----------

async def check_subscription(bot: Bot, user_id: int) -> bool:
    """Проверяет, подписан ли пользователь на канал/группу"""
    if REQUIRED_CHANNEL_ID is None:
        return True

    try:
        member = await bot.get_chat_member(chat_id=REQUIRED_CHANNEL_ID, user_id=user_id)
        return member.status in ["member", "administrator", "creator"]
    except Exception:
        return False


def subscription_keyboard():
    """Клавиатура с кнопками для подписки"""
    if REQUIRED_CHANNEL_ID is None:
        return None

    # Формирование ссылки на канал
    if isinstance(REQUIRED_CHANNEL_ID, str) and REQUIRED_CHANNEL_ID.startswith("@"):
        channel_link = f"https://t.me/{REQUIRED_CHANNEL_ID[1:]}"
    else:
        channel_link = None

    buttons = []
    if channel_link:
        buttons.append([InlineKeyboardButton(text="📢 Подписаться", url=channel_link)])
    buttons.append([InlineKeyboardButton(text="✅ Проверить подписку", callback_data="check_sub")])

    return InlineKeyboardMarkup(inline_keyboard=buttons)


class SubscriptionMiddleware(BaseMiddleware):
    """Middleware для проверки подписки на канал"""

    async def __call__(
        self,
        handler: Callable[[Update, Dict[str, Any]], Awaitable[Any]],
        event: Update,
        data: Dict[str, Any]
    ) -> Any:

        if REQUIRED_CHANNEL_ID is None:
            return await handler(event, data)

        # Получаем user_id и callback_data из события
        user_id = None
        callback_data = None

        if event.message:
            user_id = event.message.from_user.id
        elif event.callback_query:
            user_id = event.callback_query.from_user.id
            callback_data = event.callback_query.data

        if user_id is None:
            return await handler(event, data)

        # ВАЖНО: пропускаем callback "check_sub" без проверки,
        # чтобы обработчик мог сам проверить подписку
        if callback_data == "check_sub":
            return await handler(event, data)

        # Проверяем подписку
        is_subscribed = await check_subscription(data["bot"], user_id)

        if not is_subscribed:
            # Проверяем, использовал ли пользователь бота ранее
            was_active = user_id in active_users

            # Формируем ссылку на канал для текста
            channel_text = f"\n\n🔗 {CHANNEL_LINK}" if CHANNEL_LINK else ""

            if was_active:
                message_text = (
                    "❌ Вы отписались от канала!\n\n"
                    f"Для продолжения работы с ботом необходимо подписаться на наш канал.{channel_text}"
                )
                # Удаляем из активных пользователей
                active_users.discard(user_id)
                # Очищаем состояние
                await data["state"].clear()
            else:
                message_text = (
                    "👋 Добро пожаловать!\n\n"
                    f"Для использования бота необходимо подписаться на наш канал.{channel_text}"
                )

            if event.callback_query:
                await event.callback_query.answer()
                await event.callback_query.message.answer(
                    message_text,
                    reply_markup=subscription_keyboard()
                )
            elif event.message:
                await event.message.answer(
                    message_text,
                    reply_markup=subscription_keyboard()
                )

            return

        # Пользователь подписан - добавляем в активные
        active_users.add(user_id)

        return await handler(event, data)


# ---------- Клавиатуры ----------

def topics_keyboard():
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text=topic, callback_data=f"topic:{topic}")]
            for topic in QUIZ_DATA.keys()
        ]
    )


def answers_keyboard(options):
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text=opt, callback_data=f"answer:{opt}")]
            for opt in options
        ]
    )


def next_keyboard():
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="Далее", callback_data="next"),
                InlineKeyboardButton(text="Закончить", callback_data="finish")
            ],
            [InlineKeyboardButton(text="Другая тема", callback_data="change_topic")]
        ]
    )


# ---------- Бот ----------

bot = Bot(BOT_TOKEN)
dp = Dispatcher()

# Регистрируем middleware для проверки подписки
dp.update.middleware(SubscriptionMiddleware())


@dp.callback_query(F.data == "check_sub")
async def check_subscription_callback(call: CallbackQuery, state: FSMContext):
    """Обработчик проверки подписки"""
    is_subscribed = await check_subscription(call.bot, call.from_user.id)

    if is_subscribed:
        active_users.add(call.from_user.id)
        await call.message.delete()
        await call.message.answer(
            "✅ Отлично! Вы подписаны на канал.\n\nВыберите тему:",
            reply_markup=topics_keyboard()
        )
        await state.set_state(QuizState.choosing_topic)
    else:
        await call.answer(
            "Вы еще не подписались на канал",
            show_alert=True
        )


@dp.message(CommandStart())
async def start(message: Message, state: FSMContext):
    await state.clear()
    await state.set_state(QuizState.choosing_topic)

    await message.answer(
        "Выберите тему",
        reply_markup=topics_keyboard()
    )


@dp.callback_query(F.data.startswith("topic:"))
async def choose_topic(call: CallbackQuery, state: FSMContext):
    topic = call.data.split(":", 1)[1]

    questions = copy.deepcopy(QUIZ_DATA[topic])
    random.shuffle(questions)

    await state.update_data(
        topic=topic,
        questions=questions,
        index=0
    )

    await state.set_state(QuizState.answering)

    await call.message.edit_text(f"Тема выбрана: **{topic}**")
    await send_question(call.message, state)


async def send_question(message: Message, state: FSMContext):
    data = await state.get_data()
    index = data["index"]
    questions = data["questions"]

    if index >= len(questions):
        await message.answer("Вопросы закончились")
        await state.clear()
        return

    q = questions[index]

    await message.answer(
        f"{q['question']}",
        reply_markup=answers_keyboard(q["options"])
    )



@dp.callback_query(F.data.startswith("answer:"))
async def answer_question(call: CallbackQuery, state: FSMContext):
    user_answer = call.data.split(":", 1)[1]
    data = await state.get_data()

    index = data["index"]
    question = data["questions"][index]

    correct = question["correct"]

    # Формируем текст с вопросом и ответом
    question_text = f"❓ <b>Вопрос:</b>\n{question['question']}\n\n"
    user_answer_text = f"👤 <b>Ваш ответ:</b> {user_answer}\n\n"

    if user_answer == correct:
        result_text = "✅ <b>Верно!</b>"
        text = question_text + user_answer_text + result_text
    else:
        result_text = f"❌ <b>Неверно</b>\n✔️ <b>Правильный ответ:</b> {correct}"
        text = question_text + user_answer_text + result_text

    await call.message.edit_text(text, reply_markup=next_keyboard(), parse_mode="HTML")



@dp.callback_query(F.data == "next")
async def next_question(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    await state.update_data(index=data["index"] + 1)
    await send_question(call.message, state)


@dp.callback_query(F.data == "finish")
async def finish_quiz(call: CallbackQuery, state: FSMContext):
    await state.clear()
    await call.message.edit_text("Прохождение завершено. Начать заново /start")


@dp.callback_query(F.data == "change_topic")
async def change_topic(call: CallbackQuery, state: FSMContext):
    await state.clear()
    await state.set_state(QuizState.choosing_topic)
    await call.message.edit_text("Выберите тему", reply_markup=topics_keyboard())


# ---------- Запуск ----------

async def main():
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())

